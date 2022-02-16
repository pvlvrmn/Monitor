import datetime
from datetime import datetime as dt

import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

print('Считывание файла')
try:
    df = pd.read_excel('from.xlsx')
except:
    a = input('Не найден файл from.xlsx')

inv = 1

try:
    t2 = pd.read_excel('32.xlsx')
    fn = r"32.xlsx"
    wt = load_workbook(fn)
    wy = wt.active
except:
    inv = 0

wb = Workbook()
ws = wb.active
colorFill = PatternFill(start_color='FFEE1111', end_color='FFEE1111', fill_type='solid')

lost_blanks = [379386825, 392237651, 392237652, 392237653, 392237654, 392237655, 392237656,
               392237657, 392237658, 392237659, 392237660, 369815382, 369943023, 367614998,
               367614999, 370771594, 370771596, 370796787, 353569499, 353569500, 353532457]
lost_root = [387465924470]


def RS_BIK(rs, bik):
    if bik[-2:] == '.0':
        bik = bik[:-2]
    c = bik[-3:] + rs
    summ = 0
    i = 7
    for t in c:
        summ = summ + int(t) * i % 10
        if i == 7:
            i = 1
        elif i == 1:
            i = 3
        elif i == 3:
            i = 7
    if summ % 10 != 0:
        return True
    return False


def isSnils(snils):
    koef = 9
    summ = 0
    kc = 0
    for i in snils:
        summ += int(i) * koef
        koef -= 1
        if koef == 0:
            break
    kc = summ - (101 * int(summ / 101))
    if kc == int(snils[-2:]):
        return True
    if kc == 100 and int(snils[-2:]) == 0:
        return True
    return False


def isList(list):
    koef = 11
    summ = 0
    kc = 0
    list = str(list)
    for i in list:
        summ += int(i) * koef
        koef -= 1
        if koef == 0:
            break
    kc = summ - (9 * int(summ // 9))
    if kc == int(list[-1:]):
        return True
    return False


def soc_cat(days, summ):
    if days < 2:
        days = 1
    if (summ / days * 14 > 40000 * 1.15 - 2) and (summ / days * 14 < 40000 * 1.15 + 2):
        return False
    if (summ / days * 14 > 60000 * 1.15 - 2) and (summ / days * 14 < 60000 * 1.15 + 2):
        return False
    if (summ / days * 14 > 25000 * 1.15 - 2) and (summ / days * 14 < 25000 * 1.15 + 2):
        return False
    if (summ / days * 14 > 35000 * 1.15 - 2) and (summ / days * 14 < 35000 * 1.15 + 2):
        return False
    if (summ / days * 14 > 15000 * 1.15 - 2) and (summ / days * 14 < 15000 * 1.15 + 2):
        return False
    if (summ / days * 14 > 20000 * 1.15 - 2) and (summ / days * 14 < 20000 * 1.15 + 2):
        return False
    if (summ / days * 14 > 10000 * 1.15 - 2) and (summ / days * 14 < 10000 * 1.15 + 2):
        return False

    if (summ / days * 14 > 37500 * 1.15 - 2) and (summ / days * 14 < 37500 * 1.15 + 2):  # 75000
        return False
    if (summ / days * 14 > 22500 * 1.15 - 2) and (summ / days * 14 < 22500 * 1.15 + 2):  # 45000
        return False
    if (summ / days * 14 > 62500 * 1.15 - 2) and (summ / days * 14 < 62500 * 1.15 + 2):  # 125000
        return False
    if (summ / days * 14 > 70000 * 1.15 - 2) and (summ / days * 14 < 70000 * 1.15 + 2):  # 140000
        return False
    return True


def med_cat(days, summ):
    if (summ / days / 1.15 > 3879) and (summ / days / 1.15 < 3881):
        return False
    if (summ / days / 1.15 > 2429) and (summ / days / 1.15 < 2431):
        return False
    if (summ / days / 1.15 > 1214) and (summ / days / 1.15 < 1216):
        return False
    if (summ / days / 1.15 > 949) and (summ / days / 1.15 < 951):
        return False
    if (summ / days / 1.15 > 599) and (summ / days / 1.15 < 601):
        return False
    if (summ / days / 1.7 > 3879) and (summ / days / 1.7 < 3881):
        return False
    if (summ / days / 1.7 > 2429) and (summ / days / 1.7 < 2431):
        return False
    if (summ / days / 1.7 > 1214) and (summ / days / 1.7 < 1216):
        return False
    if (summ / days / 1.7 > 949) and (summ / days / 1.7 < 951):
        return False
    if (summ / days / 1.7 > 599) and (summ / days / 1.7 < 601):
        return False
    return True


def isNaN(num):  # является НеЧислом (пустое значение)
    return num != num


def isDate(a):  # является датой
    try:
        a.year
        return True
    except:
        return False


def staj(strah, nestrah):
    if isNaN(strah) or isNaN(nestrah):
        return False
    if nestrah == 0:
        return False
    if strah < 5:
        return False
    if strah >= 5 and strah < 8:
        if strah - nestrah >= 5:
            return False
    if strah > 8 and strah - nestrah > 8:
        return False
    return True


def snils(a):
    sn = ''
    if isNaN(a):
        return '0'
    sn = sn + str(int(a))
    while (len(sn) < 11):
        sn = '0' + sn
    return sn


"""
d(1 для региона / 0 для филиалов)
i[0]    - RECORD_UQ
i[1]    - Филиал 
i[1+d]  - статус
i[2+d]  - Дата/страх
i[3+d]  - Дата фонда
i[4+d]  - Дата загрузки
i[5+d]  - СНИЛС
i[6+d]  - Фамилия
i[7+d]  - Имя
i[8+d]  - Отчество
i[9+d]  - Дата рождения
i[10+d] - Пол
i[11+d] - С/Н
i[12+d] - Адрес рег
i[13+d] - Тип удостовер
i[14+d] - Пасп данные (серия и номер)
i[15+d] - ИНН
i[16+d] - Способ выплаты
i[17+d] - Номер карты
i[18+d] - Бик банка
i[19+d] - Расчетный счет
i[20+d] - рег/номер
i[21+d] - Страхователь
i[22+d] - Добровольное страхование (да/нет)
i[23+d] - ОГРН/ЛПУ (номер леч.учрежд)
i[24+d] - Название ЛПУ
i[25+d] - Номер листа нетрудоспособности
i[26+d] - Дата выдачи ЛН
i[27+d] - Тип лн (электронный,бумажный)
i[28+d] - Дубликат (да/нет)
i[29+d] - Перерасчет (да/нет)
i[30+d] - Первичный (да/нет)
i[31+d] - Пособие (тип пособия)
i[32+d] - Причина нетрудоспособности
i[33+d] - Причина перерасчета
i[34+d] - Код вида пособия
i[35+d] - Гос услуга (да/нет)
i[36+d] - Иное
i[37+d] - Причина радиации
i[38+d] - Условия исчисл
i[39+d] - Нарушение
i[40+d] - Дата нарушения режима нахождения в стационаре
i[41+d] - Фамилия И.О. больного члена семьи
i[42+d] - Возраст (больного?)
i[43+d] - Дата начала нетрудоспособности
i[44+d] - Дата окончания нетрудоспособности
i[45+d] - Стационар с
i[46+d] - Стационар по
i[47]   - Дни/н --- филиал
i[48]   - Дата - приступить к работе
i[49]   - Дата регистрации документов в МСЭ
i[50]   - Дата начала периода оплаты
i[51]   - Дата окончания периода оплаты
i[52]   - Средний заработок
i[53]   - СДЗ
i[54]   - Период простоя
i[55]   - Стаж, лет
i[56]   - Стаж, мес
i[57]   - Нестраховой стаж, лет
i[58]   - Нестраховой стаж, мес
i[59]   - Ставка
i[60]   - Районный коэф
i[61]   - Первый год расчетного периода
i[62]   - Данные для расчёта: Сумма заработка за первый год
i[63]   - Второй год расчетного периода
i[64]   - Данные для расчёта: Сумма заработка за второй год
i[65]   - Число календарных дней, учитываемых в расчетном периоде
i[66]   - Оплаченных дней
i[67]   - Начислено
i[68]   - Удержано НДФЛ
i[69]   - Удержано переплата
i[70]   - Удержано алименты
i[71]   - Удержано исполлисты
i[72]   - О/м работы
i[73]   - Источник
i[74]   - Ранние сроки беремености (да/нет)
i[75]   - Дата постановки на учет в ранние сроки беременности
i[76]   - Период
i[77]   - Дата начала отпуска
i[78]   - Дата окончания отпуска
i[79]   - Очередность
i[80]   - Уход за несколькими детьми (да/нет)
i[81]   - ФИО ребенка
i[82]   - Номер свидетельства о рождения ребенка
i[83]   - Дата свидетельства/справки о рождении
i[84]   - Дата рождения ребенка
i[85]   - Дата смерти ребенка
i[86]   - Номер справки о неполучении пособия от матери
i[87]   - Дата справки о неполучении пособия от матери
i[88]   - Номер справки о неполучении пособия от отца
i[89]   - Дата справки о неполучении пособия от отца
i[90]   - Примечание
i[91]   - Номер извещения
i[92]   - Дата извещения
"""

line = 2  # C какой строки выводить ошибки
posob = 0  # вид проверяемого пособия

d = 0
if list(df)[1] == 'Ф/Код':  # Проверка на наличие столбца кода филиала
    d = 1  # Если столбец есть то выводить его значение в файле с найденными ошибками

if inv == 1:
    for i in df.itertuples(index=False):
        if ((i[1 + d] == '005-Расч' or i[1 + d] == '007-Платим' or i[1 + d] == '074-Под' or
             i[1 + d] == '009-ИспП' or (i[1 + d] == '003-ГкР')) and i[36 + d] == 32):
            if i[5 + d] not in iter(t2['СНИЛС']):
                row = (snils(i[5 + d]), i[50].date())
                wy.append(row)
    wt.save(fn)
    wt.close()
    t2 = pd.read_excel('32.xlsx')

for i in df.itertuples(index=False):
    if (i[31 + d] == '01-В/Н') or (i[31 + d] == '05-У/Р') or (i[31 + d] == '02-Б/Р') or \
            (i[31 + d] == '03-Р/С') or (i[31 + d] == '04-Р/Р') or (i[31 + d] == '06-Н/С'):
        posob = 1
    if (i[31 + d] == '98-С/В') or (i[31 + d] == '99-С/В'):
        posob = 2
    break

if posob == 1:
    if d == 1:
        ws.cell(1, 1, 'Филиал')
        ws.column_dimensions[get_column_letter(1)].width = '10'
    ws.cell(1, 2, 'Номер записи')
    ws.column_dimensions[get_column_letter(2)].width = '15'
    ws.cell(1, 3, 'СНИЛС')  # Оглавление столбцов файла с ошибками
    ws.column_dimensions[get_column_letter(3)].width = '13'
    ws.cell(1, 4, 'Дата фонда')  # Оглавление столбцов файла с ошибками
    ws.column_dimensions[get_column_letter(4)].width = '12'
    ws.cell(1, 5, 'Пособие')  # Оглавление столбцов файла с ошибками
    ws.column_dimensions[get_column_letter(5)].width = '12'
    ws.cell(1, 6, 'ФИО получателя')  # Оглавление столбцов файла с ошибками
    ws.column_dimensions[get_column_letter(6)].width = '40'
    ws.cell(1, 7, 'Описание ошибки (ошибки выделенные зеленым принимаем для сведения)')
    ws.column_dimensions[get_column_letter(7)].width = '75'
    ws.cell(1, 8, 'Доп')
    ws.column_dimensions[get_column_letter(8)].width = '50'
    ws.cell(1, 9, 'Примечание')
    ws.column_dimensions[get_column_letter(9)].width = '60'
    print('\nПоиск ошибок')

    for i in df.itertuples(index=False):  # цикл для проверки пособий с 1 по 6

        if isNaN(i[9 + d]):
            if d == 1:
                ws.cell(line, 1, i[1])
            ws.cell(line, 2, i[0])
            ws.cell(line, 3, snils(i[5 + d]))
            ws.cell(line, 4, i[3 + d].date())
            ws.cell(line, 5, i[31 + d])
            ws.cell(line, 5, i[31 + d])
            ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
            ws.cell(line, 7, 'Не указана дата рождения')
            if not isNaN(i[90]):
                ws.cell(line, 9, str(i[90]))
            line += 1

        if isDate(i[9 + d]) and isDate(i[3 + d]):
            if ((i[3 + d].year - i[9 + d].year) > 80) or ((i[3 + d].year - i[9 + d].year) < 16):
                if d == 1:
                    ws.cell(line, 1, i[1])
                ws.cell(line, 2, i[0])
                ws.cell(line, 3, snils(i[5 + d]))
                ws.cell(line, 4, i[3 + d].date())
                ws.cell(line, 5, i[31 + d])
                ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                ws.cell(line, 7, 'Уточнить возраст.')
                ws.cell(line, 8, str(i[9 + d]))
                if not isNaN(i[90]):
                    ws.cell(line, 9, str(i[90]))
                line += 1

        if ((i[1 + d] == '005-Расч' or i[1 + d] == '007-Платим' or i[1 + d] == '074-Под' or
             i[1 + d] == '009-ИспП') and (i[29 + d] == 'Нет') and i[67] < 0):
            if d == 1:
                ws.cell(line, 1, i[1])
            ws.cell(line, 2, i[0])
            ws.cell(line, 3, snils(i[5 + d]))
            ws.cell(line, 4, i[3 + d].date())
            ws.cell(line, 5, i[31 + d])
            ws.cell(line, 5, i[31 + d])
            ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
            ws.cell(line, 7, 'Сумма меньше нуля без признака перерасчет')
            ws.cell(line, 8, str(i[67]))
            if not isNaN(i[90]):
                ws.cell(line, 9, str(i[90]))
            line += 1

        if not isNaN(i[69]) and not isNaN(i[68]):
            if i[67] - i[68] - i[69] == 0:
                if d == 1:
                    ws.cell(line, 1, i[1])
                ws.cell(line, 2, i[0])
                ws.cell(line, 3, snils(i[5 + d]))
                ws.cell(line, 4, i[3 + d].date())
                ws.cell(line, 5, i[31 + d])
                ws.cell(line, 5, i[31 + d])
                ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                ws.cell(line, 7, 'Удержанно из-за НДФЛ или переплаты столько же, сколько и начислено')
                if not isNaN(i[90]):
                    ws.cell(line, 9, str(i[90]))
                line += 1

        if i[13 + d] == 'Паспорт иностранного гражданина' and not (isNaN(i[55]) or isNaN(i[56])):
            if i[55] < 1:
                if i[56] < 6:
                    if d == 1:
                        ws.cell(line, 1, i[1])
                    ws.cell(line, 2, i[0])
                    ws.cell(line, 3, snils(i[5 + d]))
                    ws.cell(line, 4, i[3 + d].date())
                    ws.cell(line, 5, i[31 + d])
                    ws.cell(line, 5, i[31 + d])
                    ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                    ws.cell(line, 7, 'Нет стажа у иностранного гражданина')
                    ws.cell(line, 8, 'Лет: ' + str(i[55]) + '  Мес: ' + str(i[56]))
                    if not isNaN(i[90]):
                        ws.cell(line, 9, str(i[90]))
                    line += 1

        if inv == 1:
            if i[36 + d] != 32 and (i[31 + d] == '01-В/Н'):
                if str(i[38 + d]).find('45') == -1 and i[32 + d] != 3 and i[32 + d] != 11 and i[32 + d] != 9:
                    for t in t2.itertuples(index=False):
                        if i[5 + d] == t[0] and i[50] > t[1]:
                            if d == 1:
                                ws.cell(line, 1, i[1])
                            ws.cell(line, 2, i[0])
                            ws.cell(line, 3, snils(i[5 + d]))
                            ws.cell(line, 4, i[3 + d].date())
                            ws.cell(line, 5, i[31 + d])
                            ws.cell(line, 5, i[31 + d])
                            ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                            ws.cell(line, 7, 'Ранее была инвалидность, но не заполнено условие исчисления')
                            if not isNaN(i[90]):
                                ws.cell(line, 9, str(i[90]))
                            line += 1

        if not isNaN(i[2 + d]):
            if (i[3 + d].date() - i[2 + d].date()).days > 5:
                if d == 1:
                    ws.cell(line, 1, i[1])
                ws.cell(line, 2, i[0])
                ws.cell(line, 3, snils(i[5 + d]))
                ws.cell(line, 4, i[3 + d].date())
                ws.cell(line, 5, i[31 + d])
                ws.cell(line, 5, i[31 + d])
                ws.cell(line, 5, i[31 + d])
                ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                ws.cell(line, 7, 'Дата фонда позже даты страхователя более чем на 5 дней')
                ws.cell(line, 8, i[2 + d].date())
                if not isNaN(i[90]):
                    ws.cell(line, 9, str(i[90]))
                line += 1

        if ((i[1 + d] == '005-Расч' or i[1 + d] == '007-Платим' or i[1 + d] == '074-Под' or i[1 + d] == '009-ИспП')
                and i[67] == 0):
            if (d == 1):
                ws.cell(line, 1, i[1])
            ws.cell(line, 2, i[0])
            ws.cell(line, 3, snils(i[5 + d]))
            ws.cell(line, 4, i[3 + d].date())
            ws.cell(line, 5, i[31 + d])
            ws.cell(line, 5, i[31 + d])
            ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
            ws.cell(line, 7, 'Расчитанная сумма равна 0')
            if not isNaN(i[90]):
                ws.cell(line, 9, str(i[90]))
            line += 1

        if not isNaN(i[43 + d]) and not isNaN(i[44 + d]) and not isNaN(i[50]) and not isNaN(i[51]):
            if not isDate(i[44 + d]):
                if str(i[25 + d]).startswith('999', 0, 3) == (not True) and i[31 + d] == '01-В/Н' and i[30 + d] == 'Да' \
                        and (dt.strptime(i[44 + d], '%d.%m.%Y').date() - i[43 + d].date()).days == (
                        i[51].date() - i[50].date()).days \
                        and i[32 + d] != 3 and i[32 + d] != 9 and i[32 + d] != 12 and i[32 + d] != 13 and i[
                    31 + d] != 14 and i[22+d] == 'Нет':
                    if d == 1:
                        ws.cell(line, 1, i[1])
                    ws.cell(line, 2, i[0])
                    ws.cell(line, 3, snils(i[5 + d]))
                    ws.cell(line, 4, i[3 + d].date())
                    ws.cell(line, 5, i[31 + d])
                    ws.cell(line, 5, i[31 + d])
                    ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                    ws.cell(line, 7, 'Оплата первых трех дней нетрудоспособности была произведена ФСС')
                    ws.cell(line, 8, 'Добровольное страхование: '+str(i[22+d]))
                    if not isNaN(i[90]):
                        ws.cell(line, 9, str(i[90]))
                    line += 1

            if isDate(i[44 + d]):
                if str(i[25 + d]).startswith('999', 0, 3) == (not True) and i[31 + d] == '01-В/Н' and i[30 + d] == 'Да' \
                        and (i[44 + d].date() - i[43 + d].date()).days == (i[51].date() - i[50].date()).days \
                        and i[32 + d] != 3 and i[32 + d] != 9 and i[32 + d] != 12 and i[32 + d] != 13 and i[
                    31 + d] != 14:
                    if d == 1:
                        ws.cell(line, 1, i[1])
                    ws.cell(line, 2, i[0])
                    ws.cell(line, 3, snils(i[5 + d]))
                    ws.cell(line, 4, i[3 + d].date())
                    ws.cell(line, 5, i[31 + d])
                    ws.cell(line, 5, i[31 + d])
                    ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                    ws.cell(line, 7, 'Оплата первых трех дней нетрудоспособности была произведена ФСС')
                    ws.cell(line, 8, 'Добровольное страхование: '+str(i[22+d]))
                    if not isNaN(i[90]):
                        ws.cell(line, 9, str(i[90]))
                    line += 1

        if not isNaN(i[50]) and not isNaN(i[51]) and not isNaN(i[66]):
            if str(i[25 + d]).startswith('999', 0, 3) == (not True) and i[29 + d] == 'Нет' and i[30 + d] == 'Нет' \
                    and i[31 + d] == '01-В/Н' and (i[51].date() - i[50].date()).days - i[66] != -1 and i[36 + d] != 34:
                if d == 1:
                    ws.cell(line, 1, i[1])
                ws.cell(line, 2, i[0])
                ws.cell(line, 3, snils(i[5 + d]))
                ws.cell(line, 4, i[3 + d].date())
                ws.cell(line, 5, i[31 + d])
                ws.cell(line, 5, i[31 + d])
                ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                ws.cell(line, 7, 'Сформирован не весь период для оплаты!!!')
                if not isNaN(i[90]):
                    ws.cell(line, 9, str(i[90]))
                line += 1
        '''
        Гульнара Ринатовна просила убрать, бесполезная проверка
        if not isNaN(i[57]) and not isNaN(i[58]) and i[57] != 0 and i[58] != 0 or \
                isNaN(i[57]) and not isNaN(i[58]) and i[58] != 0 or \
                not isNaN(i[57]) and isNaN(i[58]) and i[57] != 0:
            if d == 1:
                ws.cell(line, 1, i[1])
            ws.cell(line, 2, i[0])
            ws.cell(line, 3, snils(i[5 + d]))
            ws.cell(line, 4, i[3 + d].date())
            ws.cell(line, 5, i[31 + d])
            ws.cell(line, 5, i[31 + d])
            ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
            ws.cell(line, 7, 'Проверить нестраховой стаж')
            if not isNaN(i[90]):
                ws.cell(line, 9, str(i[90]))
            line += 1
        '''

        if i[72] == 'Нет' and i[59] > 0.5:
            if d == 1:
                ws.cell(line, 1, i[1])
            ws.cell(line, 2, i[0])
            ws.cell(line, 3, snils(i[5 + d]))
            ws.cell(line, 4, i[3 + d].date())
            ws.cell(line, 5, i[31 + d])
            ws.cell(line, 5, i[31 + d])
            ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
            ws.cell(line, 7, 'Ставка больше 0,5 на неосновном месте работы')
            if not isNaN(i[90]):
                ws.cell(line, 9, str(i[90]))
            line += 1

        if isNaN(i[5 + d]):
            if d == 1:
                ws.cell(line, 1, i[1])
            ws.cell(line, 2, i[0])
            ws.cell(line, 4, i[3 + d].date())
            ws.cell(line, 5, i[31 + d])
            ws.cell(line, 5, i[31 + d])
            ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
            ws.cell(line, 7, 'Не заполнен СНИЛС')
            if not isNaN(i[90]):
                ws.cell(line, 9, str(i[90]))
            line += 1

        if not isNaN(i[8 + d]) and len(i[8 + d]) < 2:
            if d == 1:
                ws.cell(line, 1, i[1])
            ws.cell(line, 2, i[0])
            ws.cell(line, 3, snils(i[5 + d]))
            ws.cell(line, 4, i[3 + d].date())
            ws.cell(line, 5, i[31 + d])
            ws.cell(line, 5, i[31 + d])
            ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
            ws.cell(line, 7, 'Отчество состоит всего из 1 символа')
            if not isNaN(i[90]):
                ws.cell(line, 9, str(i[90]))
            line += 1

        if not (isNaN(i[67]) or isNaN(i[68])):
            ndfl13 = round(i[67] * 0.13, 0)
            ndfl30 = round(i[67] * 0.3, 0)
            if ndfl13 - 1 <= i[68] <= ndfl13 + 1 or ndfl30 - 1 <= i[68] <= ndfl30 + 1:
                pass
            elif ndfl13 - 5 <= i[68] <= ndfl13 + 5:
                if d == 1:
                    ws.cell(line, 1, i[1])
                ws.cell(line, 2, i[0])
                ws.cell(line, 3, snils(i[5 + d]))
                ws.cell(line, 4, i[3 + d].date())
                ws.cell(line, 5, i[31 + d])
                ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                ws.cell(line, 7,
                        'Уточнить НДФЛ: 13% от ' + str(i[67]) + ' = ' + str(ndfl13) + ', указано ' + str(i[68]))
                ws.cell(line, 8, str(i[9 + d]))
                if not isNaN(i[90]):
                    ws.cell(line, 9, str(i[90]))
                line += 1
            elif ndfl30 - 5 <= i[68] <= ndfl30 + 5:
                if d == 1:
                    ws.cell(line, 1, i[1])
                ws.cell(line, 2, i[0])
                ws.cell(line, 3, snils(i[5 + d]))
                ws.cell(line, 4, i[3 + d].date())
                ws.cell(line, 5, i[31 + d])
                ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                ws.cell(line, 7,
                        'Уточнить НДФЛ: 30% от ' + str(i[67]) + ' = ' + str(ndfl30) + ', указано ' + str(i[68]))
                ws.cell(line, 8, str(i[9 + d]))
                if not isNaN(i[90]):
                    ws.cell(line, 9, str(i[90]))
                line += 1
            else:
                if d == 1:
                    ws.cell(line, 1, i[1])
                ws.cell(line, 2, i[0])
                ws.cell(line, 3, snils(i[5 + d]))
                ws.cell(line, 4, i[3 + d].date())
                ws.cell(line, 5, i[31 + d])
                ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                ws.cell(line, 7, 'Уточнить НДФЛ, указано ' + str(i[68]))
                ws.cell(line, 8, str(i[9 + d]))
                if not isNaN(i[90]):
                    ws.cell(line, 9, str(i[90]))
                line += 1

        if i[18 + d] == 48073795:
            if d == 1:
                ws.cell(line, 1, i[1])
            ws.cell(line, 2, i[0])
            ws.cell(line, 3, snils(i[5 + d]))
            ws.cell(line, 4, i[3 + d].date())
            ws.cell(line, 5, i[31 + d])
            ws.cell(line, 5, i[31 + d])
            ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
            ws.cell(line, 7, 'БИК 048073795 более не действителен')
            if not isNaN(i[90]):
                ws.cell(line, 9, str(i[90]))
            line += 1

        if i[18 + d] == 48073001:
            if d == 1:
                ws.cell(line, 1, i[1])
            ws.cell(line, 2, i[0])
            ws.cell(line, 3, snils(i[5 + d]))
            ws.cell(line, 4, i[3 + d].date())
            ws.cell(line, 5, i[31 + d])
            ws.cell(line, 5, i[31 + d])
            ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
            ws.cell(line, 7, 'БИК 048073001 НАЦ банка')
            if not isNaN(i[90]):
                ws.cell(line, 9, str(i[90]))
            line += 1

        if i[18 + d] == 48073917:
            if d == 1:
                ws.cell(line, 1, i[1])
            ws.cell(line, 2, i[0])
            ws.cell(line, 3, snils(i[5 + d]))
            ws.cell(line, 4, i[3 + d].date())
            ws.cell(line, 5, i[31 + d])
            ws.cell(line, 5, i[31 + d])
            ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
            ws.cell(line, 7, 'БИК 048073917 более не действителен (ликвидирован)')
            if not isNaN(i[90]):
                ws.cell(line, 9, str(i[90]))
            line += 1

        if not isNaN(i[21 + d]) and not isNaN(i[6 + d]) and i[22+d] == "Нет":
            if (i[21 + d].upper().find(i[6 + d].upper(), 0, len(i[21 + d]))) != -1:
                if d == 1:
                    ws.cell(line, 1, i[1])
                ws.cell(line, 2, i[0])
                ws.cell(line, 3, snils(i[5 + d]))
                ws.cell(line, 4, i[3 + d].date())
                ws.cell(line, 5, i[31 + d])
                ws.cell(line, 5, i[31 + d])
                ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                ws.cell(line, 7, 'В наименовании страхователя содержится фамилия получателя, но нет признака добровольщика')
                ws.cell(line, 8, i[21 + d])
                if not isNaN(i[90]):
                    ws.cell(line, 9, str(i[90]))
                line += 1

        if i[16 + d] == 'К' and isNaN(i[17 + d]):
            if d == 1:
                ws.cell(line, 1, i[1])
            ws.cell(line, 2, i[0])
            ws.cell(line, 3, snils(i[5 + d]))
            ws.cell(line, 4, i[3 + d].date())
            ws.cell(line, 5, i[31 + d])
            ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
            ws.cell(line, 7, 'Указано способ перечисления Карта, а номер карты не указан')
            if not isNaN(i[90]):
                ws.cell(line, 9, str(i[90]))
            line += 1

        if i[16 + d] == 'К' and isNaN(i[17 + d]) and not isNaN(i[19 + d]):
            if d == 1:
                ws.cell(line, 1, i[1])
            ws.cell(line, 2, i[0])
            ws.cell(line, 3, snils(i[5 + d]))
            ws.cell(line, 4, i[3 + d].date())
            ws.cell(line, 5, i[31 + d])
            ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
            ws.cell(line, 7, 'Неверно указан способ перечисления')
            if not isNaN(i[90]):
                ws.cell(line, 9, str(i[90]))
            line += 1

        if not isSnils(snils(i[5 + d])):
            if d == 1:
                ws.cell(line, 1, i[1])
            ws.cell(line, 2, i[0])
            ws.cell(line, 3, snils(i[5 + d]))
            ws.cell(line, 4, i[3 + d].date())
            ws.cell(line, 5, i[31 + d])
            ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
            ws.cell(line, 7, 'СНИЛС не существует')
            if not isNaN(i[90]):
                ws.cell(line, 9, str(i[90]))
            line += 1

        if not (isNaN(i[18 + d]) or isNaN(i[19 + d])) and RS_BIK(str(i[19 + d]), str(i[18 + d])):
            if (d == 1):
                ws.cell(line, 1, i[1])
            ws.cell(line, 2, i[0])
            ws.cell(line, 3, snils(i[5 + d]))
            ws.cell(line, 4, i[3 + d].date())
            ws.cell(line, 5, i[31 + d])
            ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
            ws.cell(line, 7, 'Расчетный счет не принадлежит банку')
            if str(i[18 + d]) == '18073401':
                ws.cell(line, 8, 'БИК 18073401 принадлежит казначейству')
            if not isNaN(i[90]):
                ws.cell(line, 9, str(i[90]))
            line += 1

        if (i[16 + d] == 'К' and not isNaN(i[17 + d]) and len(str(int(i[17 + d]))) != 16):
            if (d == 1):
                ws.cell(line, 1, i[1])
            ws.cell(line, 2, i[0])
            ws.cell(line, 3, snils(i[5 + d]))
            ws.cell(line, 4, i[3 + d].date())
            ws.cell(line, 5, i[31 + d])
            ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
            ws.cell(line, 7, 'Длина номера карты должна быть 16 символов')
            if not isNaN(i[90]):
                ws.cell(line, 9, str(i[90]))
            line += 1

        if (i[16 + d] == 'К' and not isNaN(i[17 + d]) and not ((str(i[17 + d]).startswith('2') or
                                                                (str(i[17 + d]).startswith('6')) or str(
                    i[17 + d]).startswith('35')) or (str(i[17 + d]).startswith('5')))):
            if d == 1:
                ws.cell(line, 1, i[1])
            ws.cell(line, 2, i[0])
            ws.cell(line, 3, snils(i[5 + d]))
            ws.cell(line, 4, i[3 + d].date())
            ws.cell(line, 5, i[31 + d])
            ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
            ws.cell(line, 7, 'Номер карты не соответствует картам МИР')
            if not isNaN(i[90]):
                ws.cell(line, 9, str(i[90]))
            line += 1

        if i[16 + d] == 'Б' and isNaN(i[19 + d]) and not isNaN(i[17 + d]):
            if (d == 1):
                ws.cell(line, 1, i[1])
            ws.cell(line, 2, i[0])
            ws.cell(line, 3, snils(i[5 + d]))
            ws.cell(line, 4, i[3 + d].date())
            ws.cell(line, 5, i[31 + d])
            ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
            ws.cell(line, 7, 'Неверно указан способ перечисления')
            if not isNaN(i[90]):
                ws.cell(line, 9, str(i[90]))
            line += 1

        if i[16 + d] == 'Б' and isNaN(i[19 + d]) and isNaN(i[17 + d]):
            if d == 1:
                ws.cell(line, 1, i[1])
            ws.cell(line, 2, i[0])
            ws.cell(line, 3, snils(i[5 + d]))
            ws.cell(line, 4, i[3 + d].date())
            ws.cell(line, 5, i[31 + d])
            ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
            ws.cell(line, 7, 'Не указан расчетный счет')
            if not isNaN(i[90]):
                ws.cell(line, 9, str(i[90]))
            line += 1

        if i[16 + d] == 'Б' and not isNaN(i[19 + d]) and len(str(i[19 + d])) != 20:
            if d == 1:
                ws.cell(line, 1, i[1])
            ws.cell(line, 2, i[0])
            ws.cell(line, 3, snils(i[5 + d]))
            ws.cell(line, 4, i[3 + d].date())
            ws.cell(line, 5, i[31 + d])
            ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
            ws.cell(line, 7, 'Длина номера счета должна быть 20 символов')
            ws.cell(line, 8, i[19 + d])
            if not isNaN(i[90]):
                ws.cell(line, 9, str(i[90]))
            line += 1

        if i[16 + d] == 'П' and not isNaN(i[12 + d]) and len(str(i[12 + d])) > 70:
            if d == 1:
                ws.cell(line, 1, i[1])
            ws.cell(line, 2, i[0])
            ws.cell(line, 3, snils(i[5 + d]))
            ws.cell(line, 4, i[3 + d].date())
            ws.cell(line, 5, i[31 + d])
            ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
            ws.cell(line, 7, 'Длина адреса регистрации должна быть меньше 70 символов')
            ws.cell(line, 8, i[19 + d])
            if not isNaN(i[90]):
                ws.cell(line, 9, str(i[90]))
            line += 1

        if i[72] == 'Нет' and i[59] >= 1:
            if d == 1:
                ws.cell(line, 1, i[1])
            ws.cell(line, 2, i[0])
            ws.cell(line, 3, snils(i[5 + d]))
            ws.cell(line, 4, i[3 + d].date())
            ws.cell(line, 5, i[31 + d])
            ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
            ws.cell(line, 7, 'Не основное место работы, а ставка >=1')
            if not isNaN(i[90]):
                ws.cell(line, 9, str(i[90]))
            line += 1

        if (i[13 + d] == 'Паспорт иностранного гражданина' and len(
                ''.join(x for x in str(i[14 + d]) if x.isdigit())) == 10 and
                len(''.join(x for x in str(i[14 + d]) if not x.isdigit())) == 0):
            if d == 1:
                ws.cell(line, 1, i[1])
            ws.cell(line, 2, i[0])
            ws.cell(line, 3, snils(i[5 + d]))
            ws.cell(line, 4, i[3 + d].date())
            ws.cell(line, 5, i[31 + d])
            ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
            ws.cell(line, 7, 'Тип удостоверения личности соответствует паспорту гражданина РФ')
            if not isNaN(i[90]):
                ws.cell(line, 9, str(i[90]))
            line += 1

        if (i[60] != 1.15) and i[65] > 0:
            if i[52] / i[65] < 12792 * 24 / 730 * i[60]:
                if d == 1:
                    ws.cell(line, 1, i[1])
                ws.cell(line, 2, i[0])
                ws.cell(line, 3, snils(i[5 + d]))
                ws.cell(line, 4, i[3 + d].date())
                ws.cell(line, 5, i[31 + d])
                ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                ws.cell(line, 7, 'СДЗ меньше МРОТ. Проверить районный коэффициент. Не равен 1.15').fill = PatternFill(
                    start_color='008000', end_color='008000', fill_type='solid')
                ws.cell(line, 8, str(i[60]) + "  " + str(i[12 + d]))
                if not isNaN(i[90]):
                    ws.cell(line, 9, str(i[90]))
                line += 1

        if (i[31 + d] == '01-В/Н' or i[31 + d] == '02-Б/Р' or i[31 + d] == '04-Р/Р' or i[31 + d] == '05-У/Р') and not isNaN(i[19+d]):
            if i[22 + d] == 'Да' and str(i[19+d])[:6] == '408028':
                if d == 1:
                    ws.cell(line, 1, i[1])
                ws.cell(line, 2, i[0])
                ws.cell(line, 3, snils(i[5 + d]))
                ws.cell(line, 4, i[3 + d].date())
                ws.cell(line, 5, i[31 + d])
                ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                ws.cell(line, 7, 'Проверить счет добровольщика. Не может быть расчетным счетом ИП.')
                ws.cell(line, 8, str(i[19+d]))
                if not isNaN(i[90]):
                    ws.cell(line, 9, str(i[90]))
                line += 1

        if not (isNaN(i[75]) or isNaN(i[67])) and (i[31 + d] == '03-Р/С'):
            if (i[75].date() < datetime.date(2021, 2, 1)) and i[67] / i[60] > 680:
                if d == 1:
                    ws.cell(line, 1, i[1])
                ws.cell(line, 2, i[0])
                ws.cell(line, 3, snils(i[5 + d]))
                ws.cell(line, 4, i[3 + d].date())
                ws.cell(line, 5, i[31 + d])
                ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                ws.cell(line, 7, 'Старое пособие расчитано по новому постановлению')
                if not isNaN(i[90]):
                    ws.cell(line, 9, str(i[90]))
                line += 1

        if (i[31 + d] == '02-Б/Р') and not (i[66] < 141 or i[66] == 156 or i[66] == 194 or isNaN(i[66])):
            if d == 1:
                ws.cell(line, 1, i[1])
            ws.cell(line, 2, i[0])
            ws.cell(line, 3, snils(i[5 + d]))
            ws.cell(line, 4, i[3 + d].date())
            ws.cell(line, 5, i[31 + d])
            ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
            ws.cell(line, 7, 'Неверное количество оплаченных дней')
            ws.cell(line, 8, str(i[66]))
            if not isNaN(i[90]):
                ws.cell(line, 9, str(i[90]))
            line += 1

        if (i[31 + d] == '02-Б/Р') and not isNaN(i[50]) and not isNaN(i[51]) and i[29 + d] == 'Нет':
            if i[50].date() > dt.today().date() and (i[51].date() - i[50].date()).days > 100:
                if d == 1:
                    ws.cell(line, 1, i[1])
                ws.cell(line, 2, i[0])
                ws.cell(line, 3, snils(i[5 + d]))
                ws.cell(line, 4, i[3 + d].date())
                ws.cell(line, 5, i[31 + d])
                ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                ws.cell(line, 7, 'Дата начала периода оплаты еще не наступила')
                ws.cell(line, 8, i[50].date())
                if not isNaN(i[90]):
                    ws.cell(line, 9, str(i[90]))
                line += 1

        if (i[31 + d] == '05-У/Р') and not isNaN(i[50]) and not isNaN(i[51]) and i[29 + d] == 'Нет':
            if i[50].date() > dt.today().date():
                if d == 1:
                    ws.cell(line, 1, i[1])
                ws.cell(line, 2, i[0])
                ws.cell(line, 3, snils(i[5 + d]))
                ws.cell(line, 4, i[3 + d].date())
                ws.cell(line, 5, i[31 + d])
                ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                ws.cell(line, 7, 'Дата начала периода оплаты еще не наступила')
                ws.cell(line, 8, i[50].date())
                if not isNaN(i[90]):
                    ws.cell(line, 9, str(i[90]))
                line += 1

        if i[31 + d] == '03-Р/С' or i[31 + d] == '04-Р/Р':
            if i[60] != 1.15:
                if d == 1:
                    ws.cell(line, 1, i[1])
                ws.cell(line, 2, i[0])
                ws.cell(line, 3, snils(i[5 + d]))
                ws.cell(line, 4, i[3 + d].date())
                ws.cell(line, 5, i[31 + d])
                ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                ws.cell(line, 7, 'Проверить районный коэффициент. Не равен 1.15')
                ws.cell(line, 8, str(i[60]) + "  " + str(i[12 + d]))
                if not isNaN(i[90]):
                    ws.cell(line, 9, str(i[90]))
                line += 1

                if isDate(i[9 + d]) and isDate(i[3 + d]):
                    if (i[3 + d].year - i[9 + d].year) > 50:
                        if d == 1:
                            ws.cell(line, 1, i[1])
                        ws.cell(line, 2, i[0])
                        ws.cell(line, 3, snils(i[5 + d]))
                        ws.cell(line, 4, i[3 + d].date())
                        ws.cell(line, 5, i[31 + d])
                        ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                        ws.cell(line, 7, 'Уточнить возраст: ' + str(i[3 + d].year - i[9 + d].year))
                        ws.cell(line, 8, str(i[9 + d]))
                        if not isNaN(i[90]):
                            ws.cell(line, 9, str(i[90]))
                        line += 1

        if i[31 + d] == '02-Б/Р':

            if not (isNaN(i[55]) or isNaN(i[56])):
                if i[55] == 0:
                    if i[56] == 0:
                        if d == 1:
                            ws.cell(line, 1, i[1])
                        ws.cell(line, 2, i[0])
                        ws.cell(line, 3, snils(i[5 + d]))
                        ws.cell(line, 4, i[3 + d].date())
                        ws.cell(line, 5, i[31 + d])
                        ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                        ws.cell(line, 7, 'Нет стажа (Письмо РО №18-10/0218-2811 от 24.06.21)')
                        ws.cell(line, 8, 'Лет: ' + str(i[55]) + '  Мес: ' + str(i[56]))
                        if not isNaN(i[90]):
                            ws.cell(line, 9, str(i[90]))
                        line += 1

            if staj(i[55], i[57]):
                if d == 1:
                    ws.cell(line, 1, i[1])
                ws.cell(line, 2, i[0])
                ws.cell(line, 3, snils(i[5 + d]))
                ws.cell(line, 4, i[3 + d].date())
                ws.cell(line, 5, i[31 + d])
                ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                ws.cell(line, 7, 'Проверить правильность начисления пособия при наличии нестрахового стажа.')
                if not isNaN(i[90]):
                    ws.cell(line, 9, str(i[90]))
                line += 1

            if not isNaN(i[66]):
                if i[30 + d] == 'Да' and i[66] == 16:
                    if d == 1:
                        ws.cell(line, 1, i[1])
                    ws.cell(line, 2, i[0])
                    ws.cell(line, 3, snils(i[5 + d]))
                    ws.cell(line, 4, i[3 + d].date())
                    ws.cell(line, 5, i[31 + d])
                    ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                    ws.cell(line, 7, 'Проверить расчетные года')
                    ws.cell(line, 8, str(i[52]))
                    if not isNaN(i[90]):
                        ws.cell(line, 9, str(i[90]))
                    line += 1

            if isDate(i[75]):
                if i[74] == 'Да' and i[75] > i[26 + d]:
                    if d == 1:
                        ws.cell(line, 1, i[1])
                    ws.cell(line, 2, i[0])
                    ws.cell(line, 3, snils(i[5 + d]))
                    ws.cell(line, 4, i[3 + d].date())
                    ws.cell(line, 5, i[31 + d])
                    ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                    ws.cell(line, 7,
                            'Дата выдачи справки о постановке на учет в ранние '
                            'сроки беременности позже даты выдачи больничного').fill = PatternFill(start_color='008000',
                                                                                                   end_color='008000',
                                                                                                   fill_type='solid')
                    ws.cell(line, 8,
                            'Дата выдачи справки ' + str(i[75]) + ', дата выдачи больничного ' + str(i[26 + d]))
                    if not isNaN(i[90]):
                        ws.cell(line, 9, str(i[90]))
                    line += 1

            if i[27 + d] == 'ЛН':
                if (str(i[25 + d])[0]) == '3' or (str(i[25 + d])[0]) == '4':
                    if not isList(int(i[25 + d])):
                        if d == 1:
                            ws.cell(line, 1, i[1])
                        ws.cell(line, 2, i[0])
                        ws.cell(line, 3, snils(i[5 + d]))
                        ws.cell(line, 4, i[3 + d].date())
                        ws.cell(line, 5, i[31 + d])
                        ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                        ws.cell(line, 7, 'Номер ЛН не существует')
                        ws.cell(line, 8, str(int(i[25 + d])))
                        if not isNaN(i[90]):
                            ws.cell(line, 9, str(i[90]))
                        line += 1

            if (((i[27 + d] == 'ЛН') and not ((str(i[25 + d])[0]) == '3') and not ((str(i[25 + d])[0]) == '4'))
                    or ((i[27 + d] == 'ЭЛН') and not ((str(i[25 + d])[0]) == '9'))):
                if d == 1:
                    ws.cell(line, 1, i[1])
                ws.cell(line, 2, i[0])
                ws.cell(line, 3, snils(i[5 + d]))
                ws.cell(line, 4, i[3 + d].date())
                ws.cell(line, 5, i[31 + d])
                ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                ws.cell(line, 7, 'Неверный тип ЛН')
                ws.cell(line, 8, str(int(i[25 + d])) + '  ' + i[27 + d])
                if not isNaN(i[90]):
                    ws.cell(line, 9, str(i[90]))
                line += 1

            if not isNaN(i[25 + d]):
                if str(i[25 + d])[:1] == '3':
                    if int(str(int(i[25 + d]))[:9]) in lost_blanks:
                        if d == 1:
                            ws.cell(line, 1, i[1])
                        ws.cell(line, 2, i[0])
                        ws.cell(line, 3, snils(i[5 + d]))
                        ws.cell(line, 4, i[3 + d].date())
                        ws.cell(line, 5, i[31 + d])
                        ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                        ws.cell(line, 7, 'Лист нетрудоспособности был утерян и не подлежит оплате !!!')
                        ws.cell(line, 8, str(int(i[25 + d])))
                        if not isNaN(i[90]):
                            ws.cell(line, 9, str(i[90]))
                        line += 1

                    if i[25 + d] in lost_root:
                        if d == 1:
                            ws.cell(line, 1, i[1])
                        ws.cell(line, 2, i[0])
                        ws.cell(line, 3, snils(i[5 + d]))
                        ws.cell(line, 4, i[3 + d].date())
                        ws.cell(line, 5, i[31 + d])
                        ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                        ws.cell(line, 7, 'Лист нетрудоспособности был утерян и не подлежит оплате !!!')
                        ws.cell(line, 8, str(int(i[25 + d])))
                        if not isNaN(i[90]):
                            ws.cell(line, 9, str(i[90]))
                        line += 1

        if i[31 + d] == '05-У/Р':  # ошибки для ухода за ребенком до 1,5 лет

            if not (isNaN(i[55]) or isNaN(i[56])):
                if i[55] == 0:
                    if i[56] == 0:
                        if d == 1:
                            ws.cell(line, 1, i[1])
                        ws.cell(line, 2, i[0])
                        ws.cell(line, 3, snils(i[5 + d]))
                        ws.cell(line, 4, i[3 + d].date())
                        ws.cell(line, 5, i[31 + d])
                        ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                        ws.cell(line, 7, 'Нет стажа')
                        ws.cell(line, 8, 'Лет: ' + str(i[55]) + '  Мес: ' + str(i[56]))
                        if not isNaN(i[90]):
                            ws.cell(line, 9, str(i[90]))
                        line += 1

            if not isNaN(i[69]):
                if d == 1:
                    ws.cell(line, 1, i[1])
                ws.cell(line, 2, i[0])
                ws.cell(line, 3, snils(i[5 + d]))
                ws.cell(line, 4, i[3 + d].date())
                ws.cell(line, 5, i[31 + d])
                ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                ws.cell(line, 7, 'Имеется удержанная сумма по переплате (проверить в приказе)')
                ws.cell(line, 8, str(i[69]))
                if not isNaN(i[90]):
                    ws.cell(line, 9, str(i[90]))
                line += 1

            if isDate(i[9 + d]) and isDate(i[3 + d]):
                if (i[3 + d].year - i[9 + d].year) > 60:
                    if d == 1:
                        ws.cell(line, 1, i[1])
                    ws.cell(line, 2, i[0])
                    ws.cell(line, 3, snils(i[5 + d]))
                    ws.cell(line, 4, i[3 + d].date())
                    ws.cell(line, 5, i[31 + d])
                    ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                    ws.cell(line, 7, 'Уточнить возраст')
                    ws.cell(line, 8, str(i[9 + d]))
                    if not isNaN(i[90]):
                        ws.cell(line, 9, str(i[90]))
                    line += 1

            if not isNaN(i[84]) and not isNaN(i[77]) and i[84] > i[77]:
                if d == 1:
                    ws.cell(line, 1, i[1])
                ws.cell(line, 2, i[0])
                ws.cell(line, 3, snils(i[5 + d]))
                ws.cell(line, 4, i[3 + d].date())
                ws.cell(line, 5, i[31 + d])
                ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                ws.cell(line, 7, 'Дата начала отпуска ранее даты рождения ребенка')
                if not isNaN(i[90]):
                    ws.cell(line, 9, str(i[90]))
                line += 1

            if isNaN(i[86]) and isNaN(i[88]):
                if d == 1:
                    ws.cell(line, 1, i[1])
                ws.cell(line, 2, i[0])
                ws.cell(line, 3, snils(i[5 + d]))
                ws.cell(line, 4, i[3 + d].date())
                ws.cell(line, 5, i[31 + d])
                ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                ws.cell(line, 7, 'Нет справки от другого родителя')
                if not isNaN(i[90]):
                    ws.cell(line, 9, str(i[90]))
                line += 1

            if (isNaN(i[61]) or isNaN(i[63])) and not (i[31 + d] == '03-Р/С') and not (i[31 + d] == '04-Р/Р'):
                if d == 1:
                    ws.cell(line, 1, i[1])
                ws.cell(line, 2, i[0])
                ws.cell(line, 3, snils(i[5 + d]))
                ws.cell(line, 4, i[3 + d].date())
                ws.cell(line, 5, i[31 + d])
                ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                ws.cell(line, 7, 'Не заполнены расчетные года')
                if not isNaN(i[90]):
                    ws.cell(line, 9, str(i[90]))
                line += 1

            if ((i[51] - i[50]).days > 31 and (i[1 + d] == '003-ГкР' or i[1 + d] == '005-Расч' or
                                               i[1 + d] == '007-Платим' or i[1 + d] == '074-Под')):
                if d == 1:
                    ws.cell(line, 1, i[1])
                ws.cell(line, 2, i[0])
                ws.cell(line, 3, snils(i[5 + d]))
                ws.cell(line, 4, i[3 + d].date())
                ws.cell(line, 5, i[31 + d])
                ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                ws.cell(line, 7, 'Проверить период оплаты')
                if not isNaN(i[90]):
                    ws.cell(line, 9, str(i[90]))
                line += 1

            if i[13 + d] == 'Паспорт иностранного гражданина':
                if d == 1:
                    ws.cell(line, 1, i[1])
                ws.cell(line, 2, i[0])
                ws.cell(line, 3, snils(i[5 + d]))
                ws.cell(line, 4, i[3 + d].date())
                ws.cell(line, 5, i[31 + d])
                ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                ws.cell(line, 7, 'Паспорт иностранного гражданина')
                if not isNaN(i[90]):
                    ws.cell(line, 9, str(i[90]))
                line += 1

            if not isNaN(i[78]) and not isNaN(i[77]) and i[78] < i[77]:
                if d == 1:
                    ws.cell(line, 1, i[1])
                ws.cell(line, 2, i[0])
                ws.cell(line, 3, snils(i[5 + d]))
                ws.cell(line, 4, i[3 + d].date())
                ws.cell(line, 5, i[31 + d])
                ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                ws.cell(line, 7, 'Дата начала отпуска позже даты окончания отпуска')
                if not isNaN(i[90]):
                    ws.cell(line, 9, str(i[90]))
                line += 1

            if isNaN(i[78]) or isNaN(i[77]):
                if d == 1:
                    ws.cell(line, 1, i[1])
                ws.cell(line, 2, i[0])
                ws.cell(line, 3, snils(i[5 + d]))
                ws.cell(line, 4, i[3 + d].date())
                ws.cell(line, 5, i[31 + d])
                ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                ws.cell(line, 7, 'Не заполнен период отпуска')
                if not isNaN(i[90]):
                    ws.cell(line, 9, str(i[90]))
                line += 1

            if isDate(i[84]) and isDate(i[83]):
                if i[84] > i[83]:
                    if d == 1:
                        ws.cell(line, 1, i[1])
                    ws.cell(line, 2, i[0])
                    ws.cell(line, 3, snils(i[5 + d]))
                    ws.cell(line, 4, i[3 + d].date())
                    ws.cell(line, 5, i[31 + d])
                    ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                    ws.cell(line, 7, 'Дата свидетельства о рождении ранее даты рождения ребенка')
                    if not isNaN(i[90]):
                        ws.cell(line, 9, str(i[90]))
                    line += 1

            if not isDate(i[83]) or not isinstance(i[82], str):  # является ли строкой
                if d == 1:
                    ws.cell(line, 1, i[1])
                ws.cell(line, 2, i[0])
                ws.cell(line, 3, snils(i[5 + d]))
                ws.cell(line, 4, i[3 + d].date())
                ws.cell(line, 5, i[31 + d])
                ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                ws.cell(line, 7, 'Не указана дата или номер свидетельства о рождении ребенка')
                if not isNaN(i[90]):
                    ws.cell(line, 9, str(i[90]))
                line += 1

            if (isinstance(i[82], str)) and len(
                    ''.join(x for x in str(i[82]) if x.isdigit())) < 3:  # является ли строкой
                if (d == 1):
                    ws.cell(line, 1, i[1])
                ws.cell(line, 2, i[0])
                ws.cell(line, 3, snils(i[5 + d]))
                ws.cell(line, 4, i[3 + d].date())
                ws.cell(line, 5, i[31 + d])
                ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                ws.cell(line, 7, 'Некорректный номер свидетельства о рождении')
                ws.cell(line, 8, i[82])
                if not isNaN(i[90]):
                    ws.cell(line, 9, str(i[90]))
                line += 1

            if (isNaN(i[84]) or isNaN(i[81])):
                if (d == 1):
                    ws.cell(line, 1, i[1])
                ws.cell(line, 2, i[0])
                ws.cell(line, 3, snils(i[5 + d]))
                ws.cell(line, 4, i[3 + d].date())
                ws.cell(line, 5, i[31 + d])
                ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                ws.cell(line, 7, 'Не указана дата рождения или ФИО ребенка')
                if not isNaN(i[90]):
                    ws.cell(line, 9, str(i[90]))
                line += 1

            if (i[1 + d] == '075-Пров' and i[78] <= datetime.date.today() and
                    ((datetime.date.today().month - 1 > i[78].month and datetime.date.today().day >= 20) or
                     (datetime.date.today().month - 2 > i[78].month and datetime.date.today().day < 20))):
                if d == 1:
                    ws.cell(line, 1, i[1])
                ws.cell(line, 2, i[0])
                ws.cell(line, 3, snils(i[5 + d]))
                ws.cell(line, 4, i[3 + d].date())
                ws.cell(line, 5, i[31 + d])
                ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                ws.cell(line, 7, 'Закрыть проверенный документ')
                if not isNaN(i[90]):
                    ws.cell(line, 9, str(i[90]))
                line += 1

            if (i[78] - i[84]).days > 550:
                if d == 1:
                    ws.cell(line, 1, i[1])
                ws.cell(line, 2, i[0])
                ws.cell(line, 3, snils(i[5 + d]))
                ws.cell(line, 4, i[3 + d].date())
                ws.cell(line, 5, i[31 + d])
                ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                ws.cell(line, 7, 'Между датой рождения ребенка и окончанием отпуска > 550 дней')
                if not isNaN(i[90]):
                    ws.cell(line, 9, str(i[90]))
                line += 1

        if (i[31 + d] == '01-В/Н') or (i[31 + d] == '06-Н/С'):  # ошибки для временной нетрудоспособности

            if not isNaN(i[43 + d]) and not isNaN(i[50]) and not isNaN(i[51]):
                if i[32 + d] == 9 and i[43 + d].date() != i[50].date() and i[22 + d] == 'Нет':
                    if d == 1:
                        ws.cell(line, 1, i[1])
                    ws.cell(line, 2, i[0])
                    ws.cell(line, 3, snils(i[5 + d]))
                    ws.cell(line, 4, i[3 + d].date())
                    ws.cell(line, 5, i[31 + d])
                    ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                    ws.cell(line, 7,
                            'Оплата первых трех дней нетрудоспособности была произведена не ФСС').fill = PatternFill(
                        start_color='008000', end_color='008000', fill_type='solid')
                    if not isNaN(i[90]):
                        ws.cell(line, 9, str(i[90]))
                    line += 1

            if i[32 + d] == 6:
                if d == 1:
                    ws.cell(line, 1, i[1])
                ws.cell(line, 2, i[0])
                ws.cell(line, 3, snils(i[5 + d]))
                ws.cell(line, 4, i[3 + d].date())
                ws.cell(line, 5, i[31 + d])
                ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                ws.cell(line, 7, 'В РБ по данному пособию выплаты не производятся')
                ws.cell(line, 8, str(i[9 + d]))
                if not isNaN(i[90]):
                    ws.cell(line, 9, str(i[90]))
                line += 1

            if isDate(i[26 + d]):
                if i[26 + d].year == i[61] or i[26 + d].year == i[63]:
                    if d == 1:
                        ws.cell(line, 1, i[1])
                    ws.cell(line, 2, i[0])
                    ws.cell(line, 3, snils(i[5 + d]))
                    ws.cell(line, 4, i[3 + d].date())
                    ws.cell(line, 5, i[31 + d])
                    ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                    ws.cell(line, 7, 'Расчетный год совпадает с тем годом в котором выдан ЛН')
                    if not isNaN(i[90]):
                        ws.cell(line, 9, str(i[90]))
                    line += 1

            if not isNaN(i[25 + d]):
                if str(i[25 + d])[:1] == '3':
                    if int(str(int(i[25 + d]))[:9]) in lost_blanks:
                        if (d == 1):
                            ws.cell(line, 1, i[1])
                        ws.cell(line, 2, i[0])
                        ws.cell(line, 3, snils(i[5 + d]))
                        ws.cell(line, 4, i[3 + d].date())
                        ws.cell(line, 5, i[31 + d])
                        ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                        ws.cell(line, 7, 'Лист нетрудоспособности был утерян и не подлежит оплате !!!')
                        ws.cell(line, 8, str(int(i[25 + d])))
                        if not isNaN(i[90]):
                            ws.cell(line, 9, str(i[90]))
                        line += 1

                    if i[25 + d] in lost_root:
                        if (d == 1):
                            ws.cell(line, 1, i[1])
                        ws.cell(line, 2, i[0])
                        ws.cell(line, 3, snils(i[5 + d]))
                        ws.cell(line, 4, i[3 + d].date())
                        ws.cell(line, 5, i[31 + d])
                        ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                        ws.cell(line, 7, 'Лист нетрудоспособности был утерян и не подлежит оплате !!!')
                        ws.cell(line, 8, str(int(i[25 + d])))
                        if not isNaN(i[90]):
                            ws.cell(line, 9, str(i[90]))
                        line += 1

            if i[32 + d] == 3 and not isNaN(i[41 + d]) and not isNaN(i[42 + d]):
                if i[42 + d] > 6:
                    if d == 1:
                        ws.cell(line, 1, i[1])
                    ws.cell(line, 2, i[0])
                    ws.cell(line, 3, snils(i[5 + d]))
                    ws.cell(line, 4, i[3 + d].date())
                    ws.cell(line, 5, i[31 + d])
                    ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                    ws.cell(line, 7, 'Карантин для больного члена семьи старше 7 лет')
                    ws.cell(line, 8, str(i[42 + d]))
                    if not isNaN(i[90]):
                        ws.cell(line, 9, str(i[90]))
                    line += 1

            if (i[27 + d] == 'ЛН') and ((str(i[25 + d])[0]) == '3'):
                if not isList(int(i[25 + d])):
                    if d == 1:
                        ws.cell(line, 1, i[1])
                    ws.cell(line, 2, i[0])
                    ws.cell(line, 3, snils(i[5 + d]))
                    ws.cell(line, 4, i[3 + d].date())
                    ws.cell(line, 5, i[31 + d])
                    ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                    ws.cell(line, 7, 'Номер ЛН не существует')
                    ws.cell(line, 8, str(int(i[25 + d])))
                    if not isNaN(i[90]):
                        ws.cell(line, 9, str(i[90]))
                    line += 1

            if (((i[27 + d] == 'ЛН') and not ((str(i[25 + d])[0]) == '3') and not ((str(i[25 + d])[0]) == '4'))
                    or ((i[27 + d] == 'ЭЛН') and not ((str(i[25 + d])[0]) == '9'))):
                if d == 1:
                    ws.cell(line, 1, i[1])
                ws.cell(line, 2, i[0])
                ws.cell(line, 3, snils(i[5 + d]))
                ws.cell(line, 4, i[3 + d].date())
                ws.cell(line, 5, i[31 + d])
                ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                ws.cell(line, 7, 'Неверный тип ЛН')
                ws.cell(line, 8, str(i[25 + d]) + '  ' + i[27 + d])
                if not isNaN(i[90]):
                    ws.cell(line, 9, str(i[90]))
                line += 1

            if i[32 + d] == 9 and (i[42 + d] == 0 or i[42 + d] == 1):
                if d == 1:
                    ws.cell(line, 1, i[1])
                ws.cell(line, 2, i[0])
                ws.cell(line, 3, snils(i[5 + d]))
                ws.cell(line, 4, i[3 + d].date())
                ws.cell(line, 5, i[31 + d])
                ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                ws.cell(line, 7, 'Возраст больного члена семьи меньше 2 лет. Проверить УР и полную ставку')
                ws.cell(line, 8, 'Возраст ' + str(int(i[42 + d])) + ', ставка ' + str(int(i[59])))
                if not isNaN(i[90]):
                    ws.cell(line, 9, str(i[90]))
                line += 1

            if not isNaN(i[26 + d]) and not isNaN(i[44 + d]) and isNaN(i[46 + d]):
                if i[26 + d] == i[44 + d] and i[28 + d] == 'Нет':
                    if d == 1:
                        ws.cell(line, 1, i[1])
                    ws.cell(line, 2, i[0])
                    ws.cell(line, 3, snils(i[5 + d]))
                    ws.cell(line, 4, i[3 + d].date())
                    ws.cell(line, 5, i[31 + d])
                    ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                    ws.cell(line, 7,
                            'Дата выдачи ЛН равна окончанию периода нетрудоспособности, но не заполнен стационар')
                    if not isNaN(i[90]):
                        ws.cell(line, 9, str(i[90]))
                    line += 1

            if staj(i[55], i[57]):
                if d == 1:
                    ws.cell(line, 1, i[1])
                ws.cell(line, 2, i[0])
                ws.cell(line, 3, snils(i[5 + d]))
                ws.cell(line, 4, i[3 + d].date())
                ws.cell(line, 5, i[31 + d])
                ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                ws.cell(line, 7, 'Проверить правильность начисления пособия при наличии нестрахового стажа.')
                if not isNaN(i[90]):
                    ws.cell(line, 9, str(i[90]))
                line += 1

            if i[36 + d] == 32 and isNaN(i[49]):
                if d == 1:
                    ws.cell(line, 1, i[1])
                ws.cell(line, 2, i[0])
                ws.cell(line, 3, snils(i[5 + d]))
                ws.cell(line, 4, i[3 + d].date())
                ws.cell(line, 5, i[31 + d])
                ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                ws.cell(line, 7, 'Установлена инвалидность и нет даты МСЭ')
                if not isNaN(i[90]):
                    ws.cell(line, 9, str(i[90]))
                line += 1

            if i[13 + d] != 'Паспорт гражданина Российской Федерации':
                if d == 1:
                    ws.cell(line, 1, i[1])
                ws.cell(line, 2, i[0])
                ws.cell(line, 3, snils(i[5 + d]))
                ws.cell(line, 4, i[3 + d].date())
                ws.cell(line, 5, i[31 + d])
                ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                ws.cell(line, 7, "Проверить соответствие документа удостоверяющего личность")
                ws.cell(line, 8, str(i[13 + d]))
                if not isNaN(i[90]):
                    ws.cell(line, 9, str(i[90]))
                line += 1

            if (i[11 + d] == 0) or (i[11 + d] == 2):
                if (d == 1):
                    ws.cell(line, 1, i[1])
                ws.cell(line, 2, i[0])
                ws.cell(line, 3, snils(i[5 + d]))
                ws.cell(line, 4, i[3 + d].date())
                ws.cell(line, 5, i[31 + d])
                ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                ws.cell(line, 7, 'Не резидент')
                if not isNaN(i[90]):
                    ws.cell(line, 9, str(i[90]))
                line += 1

            if not isNaN(i[67]) and not isNaN(i[68]) and i[67] != 0 and i[68] != 0:
                if i[31 + d] == "01-В/Н" and i[11 + d] == 1 and i[68] / i[67] > 0.28:
                    if d == 1:
                        ws.cell(line, 1, i[1])
                    ws.cell(line, 2, i[0])
                    ws.cell(line, 3, snils(i[5 + d]))
                    ws.cell(line, 4, i[3 + d].date())
                    ws.cell(line, 5, i[31 + d])
                    ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                    ws.cell(line, 7, 'Резиденту назначен налог в 30% как не резиденту')
                    if not isNaN(i[90]):
                        ws.cell(line, 9, str(i[90]))
                    line += 1

            if isNaN(i[43 + d] or isNaN(i[44 + d] and not isDate(i[44 + d]))):
                if d == 1:
                    ws.cell(line, 1, i[1])
                ws.cell(line, 2, i[0])
                ws.cell(line, 3, snils(i[5 + d]))
                ws.cell(line, 4, i[3 + d].date())
                ws.cell(line, 5, i[31 + d])
                ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                ws.cell(line, 7, 'Не заполнена дата начала или окончания периода нетрудоспособности')
                if not isNaN(i[90]):
                    ws.cell(line, 9, str(i[90]))
                line += 1

            if not isNaN(i[43 + d]) and not isNaN(i[44 + d]) and not isNaN(i[50]) and not isNaN(i[51]):
                if (dt.strptime(i[44 + d], '%d.%m.%Y').date() - i[43 + d].date()).days < (
                        i[51].date() - i[50].date()).days:
                    if (d == 1):
                        ws.cell(line, 1, i[1])
                    ws.cell(line, 2, i[0])
                    ws.cell(line, 3, snils(i[5 + d]))
                    ws.cell(line, 4, i[3 + d].date())
                    ws.cell(line, 5, i[31 + d])
                    ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                    ws.cell(line, 7, 'Период нетрудоспособности меньше периода оплаты')
                    if not isNaN(i[90]):
                        ws.cell(line, 9, str(i[90]))
                    line += 1

            if (i[59] < 1 and isNaN(i[38 + d])):
                if (d == 1):
                    ws.cell(line, 1, i[1])
                ws.cell(line, 2, i[0])
                ws.cell(line, 3, snils(i[5 + d]))
                ws.cell(line, 4, i[3 + d].date())
                ws.cell(line, 5, i[31 + d])
                ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                ws.cell(line, 7, 'Если ставка <1, то должно быть заполнено условия исчисления')
                if not isNaN(i[90]):
                    ws.cell(line, 9, str(i[90]))
                line += 1

            if not isNaN(i[43 + d]) and not isNaN(i[44 + d]):
                if isDate(i[44 + d]):
                    if i[43 + d] > i[44 + d]:
                        if d == 1:
                            ws.cell(line, 1, i[1])
                        ws.cell(line, 2, i[0])
                        ws.cell(line, 3, snils(i[5 + d]))
                        ws.cell(line, 4, i[3 + d].date())
                        ws.cell(line, 5, i[31 + d])
                        ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                        ws.cell(line, 7,
                                'Дата начала периода нетрудоспособности позже даты окончания периода нетрудоспособности')
                        if not isNaN(i[90]):
                            ws.cell(line, 9, str(i[90]))
                        line += 1

                if not isDate(i[44 + d]):
                    if i[43 + d] > dt.strptime(i[44 + d], '%d.%m.%Y'):
                        if d == 1:
                            ws.cell(line, 1, i[1])
                        ws.cell(line, 2, i[0])
                        ws.cell(line, 3, snils(i[5 + d]))
                        ws.cell(line, 4, i[3 + d].date())
                        ws.cell(line, 5, i[31 + d])
                        ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                        ws.cell(line, 7,
                                'Дата начала периода нетрудоспособности позже даты окончания периода нетрудоспособности')
                        if not isNaN(i[90]):
                            ws.cell(line, 9, str(i[90]))
                        line += 1

            if not isNaN(i[43 + d]) and not isNaN(i[3 + d]):
                if i[43 + d] > i[3 + d]:
                    if d == 1:
                        ws.cell(line, 1, i[1])
                    ws.cell(line, 2, i[0])
                    ws.cell(line, 3, snils(i[5 + d]))
                    ws.cell(line, 4, i[3 + d].date())
                    ws.cell(line, 5, i[31 + d])
                    ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                    ws.cell(line, 7, 'Дата начала периода нетрудоспособности позже даты фонда')
                    if not isNaN(i[90]):
                        ws.cell(line, 9, str(i[90]))
                    line += 1

            if not isNaN(i[66]) and not isNaN(i[41 + d]):
                if i[29 + d] == 'Нет':
                    if ((i[51].date() - i[50].date()).days) + 1 != (i[66]):
                        if (d == 1):
                            ws.cell(line, 1, i[1])
                        ws.cell(line, 2, i[0])
                        ws.cell(line, 3, snils(i[5 + d]))
                        ws.cell(line, 4, i[3 + d].date())
                        ws.cell(line, 5, i[31 + d])
                        ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                        ws.cell(line, 7, 'Период оплаты не соответствует кол-ву оплаченных дней')
                        ws.cell(line, 8, 'Причина нетрудоспособности ' + str('{0:g}'.format(i[32 + d])))
                        if not isNaN(i[90]):
                            ws.cell(line, 9, str(i[90]))
                        line += 1

            if not isNaN(i[41 + d]):
                if not (i[32 + d] == 12 or i[32 + d] == 13 or i[32 + d] == 14 or
                        i[32 + d] == 15 or i[32 + d] == 9 or i[32 + d] == 3):
                    if d == 1:
                        ws.cell(line, 1, i[1])
                    ws.cell(line, 2, i[0])
                    ws.cell(line, 3, snils(i[5 + d]))
                    ws.cell(line, 4, i[3 + d].date())
                    ws.cell(line, 5, i[31 + d])
                    ws.cell(line, 6, str(i[6 + d]) + ' ' + str(i[7 + d]) + ' ' + str(i[8 + d]))
                    ws.cell(line, 7, 'Имеется ФИО больного члена семьи, проверить код причины нетрудоспособности')
                    ws.cell(line, 8, i[32 + d])
                    if not isNaN(i[90]):
                        ws.cell(line, 9, str(i[90]))
                    line += 1

if posob == 2:  # Соц и Мед выплаты
    if d == 1:
        ws.cell(1, 1, 'Филиал')
        ws.column_dimensions[get_column_letter(1)].width = '10'
    ws.cell(1, 2, 'Номер записи')
    ws.column_dimensions[get_column_letter(2)].width = '20'
    ws.cell(1, 3, 'СНИЛС')  # Оглавление столбцов файла с ошибками
    ws.column_dimensions[get_column_letter(3)].width = '20'
    ws.cell(1, 4, 'Статус документа')
    ws.column_dimensions[get_column_letter(4)].width = '20'
    ws.cell(1, 5, 'Описание ошибки')
    ws.column_dimensions[get_column_letter(5)].width = '40'
    print('\nПоиск ошибок')
    for i in df.itertuples(index=False):

        if isNaN(i[5 + d]):
            if d == 1:
                ws.cell(line, 1, i[1])
            ws.cell(line, 2, i[0])
            ws.cell(line, 4, i[1 + d])
            ws.cell(line, 5, 'Не заполнен СНИЛС')
            line += 1

        if not isSnils(snils(i[5 + d])):
            if d == 1:
                ws.cell(line, 1, i[1])
            ws.cell(line, 2, i[0])
            ws.cell(line, 3, snils(i[5 + d]))
            ws.cell(line, 4, i[1 + d])
            ws.cell(line, 5, 'СНИЛС не существует')
            line += 1

        if isNaN(i[6 + d]):
            if d == 1:
                ws.cell(line, 1, i[1])
            ws.cell(line, 2, i[0])
            ws.cell(line, 3, snils(i[5 + d]))
            ws.cell(line, 4, i[1 + d])
            ws.cell(line, 5, 'Не заполнено Имя')
            line += 1

        if isNaN(i[7 + d]):
            if d == 1:
                ws.cell(line, 1, i[1])
            ws.cell(line, 2, i[0])
            ws.cell(line, 3, snils(i[5 + d]))
            ws.cell(line, 4, i[1 + d])
            ws.cell(line, 5, 'Не заполнена Фамилия')
            line += 1

        if isNaN(i[8 + d]):
            if d == 1:
                ws.cell(line, 1, i[1])
            ws.cell(line, 2, i[0])
            ws.cell(line, 3, snils(i[5 + d]))
            ws.cell(line, 4, i[1 + d])
            ws.cell(line, 5, 'Не заполнено Отчество')
            line += 1

        if not isNaN(i[8 + d]) and len(i[8 + d]) < 2:
            if d == 1:
                ws.cell(line, 1, i[1])
            ws.cell(line, 2, i[0])
            ws.cell(line, 3, snils(i[5 + d]))
            ws.cell(line, 4, i[1 + d])
            ws.cell(line, 5, 'Отчество состоит всего из 1 символа')
            line += 1

        if i[16 + d] == 'К' and isNaN(i[17 + d]) and isNaN(i[19 + d]):
            if d == 1:
                ws.cell(line, 1, i[1])
            ws.cell(line, 2, i[0])
            ws.cell(line, 3, snils(i[5 + d]))
            ws.cell(line, 4, i[1 + d])
            ws.cell(line, 5, 'Указано способ перечисления Карта, а номер карты не указан')
            line += 1

        if not (isNaN(i[18 + d]) or isNaN(i[19 + d])) and RS_BIK(str(i[19 + d]), str(i[18 + d])):
            if d == 1:
                ws.cell(line, 1, i[1])
            ws.cell(line, 2, i[0])
            ws.cell(line, 3, snils(i[5 + d]))
            ws.cell(line, 4, i[1 + d])
            ws.cell(line, 5, 'Расчетный счет не принадлежит банку')
            if str(i[18 + d]) == '18073401':
                ws.cell(line, 6, 'БИК 18073401 принадлежит казначейству')
            line += 1

        if i[18 + d] == 48073795:
            if d == 1:
                ws.cell(line, 1, i[1])
            ws.cell(line, 2, i[0])
            ws.cell(line, 3, snils(i[5 + d]))
            ws.cell(line, 4, i[3 + d].date())
            ws.cell(line, 5, i[31 + d])
            ws.cell(line, 5, 'БИК 048073795 более не действителен')
            line += 1

        if i[18 + d] == 48073001:
            if (d == 1):
                ws.cell(line, 1, i[1])
            ws.cell(line, 2, i[0])
            ws.cell(line, 3, snils(i[5 + d]))
            ws.cell(line, 4, i[3 + d].date())
            ws.cell(line, 5, i[31 + d])
            ws.cell(line, 5, 'БИК 048073001 более не действителен')
            line += 1

        if i[18 + d] == 48073917:
            if d == 1:
                ws.cell(line, 1, i[1])
            ws.cell(line, 2, i[0])
            ws.cell(line, 3, snils(i[5 + d]))
            ws.cell(line, 4, i[3 + d].date())
            ws.cell(line, 5, i[31 + d])
            ws.cell(line, 5, 'БИК 048073795 более не действителен (ликвидирован)')
            line += 1

        if i[16 + d] == 'Б' and isNaN(i[19 + d]) and not isNaN(i[17 + d]):
            if d == 1:
                ws.cell(line, 1, i[1])
            ws.cell(line, 2, i[0])
            ws.cell(line, 3, snils(i[5 + d]))
            ws.cell(line, 4, i[1 + d])
            ws.cell(line, 5, 'Неверно указан способ перечисления')
            line += 1

        if i[16 + d] == 'К' and not isNaN(i[17 + d]) and len(str(int(i[17 + d]))) != 16:
            if d == 1:
                ws.cell(line, 1, i[1])
            ws.cell(line, 2, i[0])
            ws.cell(line, 3, snils(i[5 + d]))
            ws.cell(line, 4, i[1 + d])
            ws.cell(line, 5, 'Длина номера карты должна быть 16 символов')
            line += 1

        if (i[16 + d] == 'К' and not isNaN(i[17 + d]) and not
        ((str(i[17 + d]).startswith('2')) or (str(i[17 + d]).startswith('6')) or
         (str(i[17 + d]).startswith('35')) or (str(i[17 + d]).startswith('5')))):
            if d == 1:
                ws.cell(line, 1, i[1])
            ws.cell(line, 2, i[0])
            ws.cell(line, 3, snils(i[5 + d]))
            ws.cell(line, 4, i[1 + d])
            ws.cell(line, 5, 'Номер карты не соответствует картам МИР')
            line += 1

        if (i[16 + d] == 'Б' and not isNaN(i[19 + d]) and not
        (str(i[19 + d]).startswith('408') or str(i[19 + d]).startswith('423') or
         str(i[19 + d]).startswith('302') or str(i[19 + d]).startswith('303'))):
            if d == 1:
                ws.cell(line, 1, i[1])
            ws.cell(line, 2, i[0])
            ws.cell(line, 3, snils(i[5 + d]))
            ws.cell(line, 4, i[1 + d])
            ws.cell(line, 5, 'Номер счета должен начинаться на 408, 423, 302 либо 303')
            line += 1

        if i[16 + d] == 'Б' and isNaN(i[19 + d]) and isNaN(i[17 + d]):
            if d == 1:
                ws.cell(line, 1, i[1])
            ws.cell(line, 2, i[0])
            ws.cell(line, 3, snils(i[5 + d]))
            ws.cell(line, 4, i[1 + d])
            ws.cell(line, 5, 'Не указан расчетный счет')
            line += 1

        if i[16 + d] == 'Б' and not isNaN(i[19 + d]) and len(str(i[19 + d])) != 20:
            if d == 1:
                ws.cell(line, 1, i[1])
            ws.cell(line, 2, i[0])
            ws.cell(line, 3, snils(i[5 + d]))
            ws.cell(line, 4, i[1 + d])
            ws.cell(line, 5, 'Длина номера счета должна быть 20 символов')
            line += 1

        if not isNaN(i[58]):
            if d == 1:
                ws.cell(line, 1, i[1])
            ws.cell(line, 2, i[0])
            ws.cell(line, 3, snils(i[5 + d]))
            ws.cell(line, 4, i[1 + d])
            ws.cell(line, 5, 'Указан повышающий коэффициент')
            line += 1

        if not isNaN(i[18 + d]):
            if len(str(int(i[18 + d]))) != 8 and len(str(int(i[18 + d]))) != 9:
                if d == 1:
                    ws.cell(line, 1, i[1])
                ws.cell(line, 2, i[0])
                ws.cell(line, 3, snils(i[5 + d]))
                ws.cell(line, 4, i[1 + d])
                ws.cell(line, 5, 'Неверно указан БИК')
                line += 1

        if isDate(i[9 + d]) and isDate(i[3 + d]):
            if (i[3 + d].year - i[9 + d].year) < 16:
                if d == 1:
                    ws.cell(line, 1, i[1])
                ws.cell(line, 2, i[0])
                ws.cell(line, 3, snils(i[5 + d]))
                ws.cell(line, 4, i[1 + d])
                ws.cell(line, 5, 'Возраст менее 16 лет. Проверить дату рождения')
                line += 1

        if (i[31 + d] == '98-С/В') and i[50] < datetime.date(2020, 11, 16):
            if d == 1:
                ws.cell(line, 1, i[1])
            ws.cell(line, 2, i[0])
            ws.cell(line, 3, snils(i[5 + d]))
            ws.cell(line, 4, i[1 + d])
            ws.cell(line, 5, 'Дата начала выплаты пособия раньше, чем 16.11.2020')
            line += 1

        if (i[31 + d] == '99-С/В') and i[50] < datetime.date(2020, 11, 1):
            if d == 1:
                ws.cell(line, 1, i[1])
            ws.cell(line, 2, i[0])
            ws.cell(line, 3, snils(i[5 + d]))
            ws.cell(line, 4, i[1 + d])
            ws.cell(line, 5, 'Дата начала выплаты пособия раньше, чем 01.11.2020')
            line += 1

        if i[60] != 1.15 and i[60] != 1.7:
            if d == 1:
                ws.cell(line, 1, i[1])
            ws.cell(line, 2, i[0])
            ws.cell(line, 3, snils(i[5 + d]))
            ws.cell(line, 4, i[1 + d])
            ws.cell(line, 5, 'Проверить районный коэффициент, не равен 1.15')
            line += 1

        if not isNaN(i[66]):
            if i[66] > 44:
                if d == 1:
                    ws.cell(line, 1, i[1])
                ws.cell(line, 2, i[0])
                ws.cell(line, 3, snils(i[5 + d]))
                ws.cell(line, 4, i[1 + d])
                ws.cell(line, 5, 'Количество оплаченных дней больше 44')
                line += 1

        if ((i[1 + d] == '005-Расч') or (i[1 + d] == '074-Под') or (i[1 + d] == '007-Платим')
                or (i[1 + d] == '009-ИспП')):
            if i[29 + d] != 'Да':
                if i[31 + d] == '98-С/В':
                    if soc_cat(i[66], i[67]) and soc_cat(i[66], i[67] / 2):
                        if d == 1:
                            ws.cell(line, 1, i[1])
                        ws.cell(line, 2, i[0])
                        ws.cell(line, 3, snils(i[5 + d]))
                        ws.cell(line, 4, i[1 + d])
                        ws.cell(line, 5, 'Начисленная сумма не соответствует ни одной из категорий работников')
                        line += 1
                else:
                    if med_cat(i[66], i[67]) and med_cat(i[66] * 2, i[67]):
                        if d == 1:
                            ws.cell(line, 1, i[1])
                        ws.cell(line, 2, i[0])
                        ws.cell(line, 3, snils(i[5 + d]))
                        ws.cell(line, 4, i[1 + d])
                        ws.cell(line, 5, 'Начисленная сумма не соответствует ни одной из категорий работников')
                        line += 1
                    if round(i[66] * 100) % 10 != 0:
                        if d == 1:
                            ws.cell(line, 1, i[1])
                        ws.cell(line, 2, i[0])
                        ws.cell(line, 3, snils(i[5 + d]))
                        ws.cell(line, 4, i[1 + d])
                        ws.cell(line, 5, 'Неверно указано количество смен')
                        line += 1

if not ws.cell(2, 2).value:
    a = input('Ошибки не найдены\n')
else:
    ws.auto_filter.ref = ws.dimensions
    wb.save('Report.xlsx')
    a = input('Отчет сформирован\n')
